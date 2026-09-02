#!/usr/bin/env python3
"""
Assemble the presale competitor-analysis workbook from the JSON files the
other scripts in this skill produced, plus your own classification
judgment (see classification.json format below).

Usage:
    python3 build_report.py \
        --serp serp_results.json \
        --classification classification.json \
        --metrics metrics.json \
        --client motors.com.ua \
        --out seo_presale.xlsx

--serp and --classification are optional — pass only the pieces you
actually ran. --metrics and --out are required (the metrics sheet is the
core deliverable).

classification.json is a JSON list you write by hand after reading the
output of ranked_keywords.py — this step needs human/model judgment, not
another API call:
    [
      {"domain": "engineparts.com.ua", "top10_count": 3,
       "category": "Тільки мотори",
       "comment": "Ключі — майже виключно коди двигунів"},
      {"domain": "motopoland.com.ua", "top10_count": 5,
       "category": "Всі автозапчастини",
       "comment": "Диски, тахографи, ГБЦ — двигун лише одна з категорій"},
      {"domain": "mototeam.com.ua", "top10_count": 1,
       "category": "ВИКЛЮЧЕНО", "comment": "Мототехніка, не авто"}
    ]
Any row whose category contains "ВИКЛЮЧЕНО" is highlighted red on the
classification sheet and excluded from the metrics sheet automatically —
you don't need to filter --metrics's input domain list yourself, just make
sure the domain names match between files.

IMPORTANT — about formulas in this workbook:
This script writes the summary rows (average / max across competitors) as
plain pre-computed numbers, not live spreadsheet formulas. That's a
deliberate workaround, not laziness: the xlsx skill's LibreOffice-based
recalculation step is required to ship any workbook that DOES use
formulas (openpyxl writes formulas with no cached value, so nothing can
read them back until they're recalculated), and in this remote sandboxed
environment that headless-LibreOffice recalc reliably hung for 400+
seconds on even a two-formula file. If you're running somewhere recalc
works fine, feel free to swap these for real =AVERAGE()/=MAX() formulas —
just budget time to run scripts/recalc.py from the xlsx skill afterwards
and confirm it actually returns before you ship the file.
"""
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="305496")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUB_FONT = Font(name=FONT, italic=True, size=9, color="666666")
NORMAL_FONT = Font(name=FONT, size=10)
BOLD_FONT = Font(name=FONT, bold=True, size=10)
CLIENT_FILL = PatternFill("solid", fgColor="FFF2CC")
EXCLUDED_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_serp_sheet(wb, serp_rows):
    ws = wb.create_sheet("ТОП-10 видача")
    ws["A1"] = "ТОП-10 видача Google"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Джерело: DataForSEO SERP API"
    ws["A2"].font = SUB_FONT

    headers = ["Запит", "Позиція", "Домен", "Заголовок", "URL"]
    hr = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers))

    r = hr + 1
    for row in serp_rows:
        ws.cell(row=r, column=1, value=row["query"]).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=row["position"]).font = NORMAL_FONT
        ws.cell(row=r, column=3, value=row["domain"]).font = NORMAL_FONT
        ws.cell(row=r, column=4, value=row["title"]).font = NORMAL_FONT
        ws.cell(row=r, column=5, value=row["url"]).font = NORMAL_FONT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        r += 1
    autofit(ws, [24, 9, 22, 45, 45])
    ws.freeze_panes = "A5"


def add_classification_sheet(wb, classification, client_domain):
    ws = wb.create_sheet("Конкуренти")
    ws["A1"] = "Список конкурентів та їх спеціалізація"
    ws["A1"].font = TITLE_FONT

    headers = ["Домен", "Разів у ТОП-10", "Асортимент", "Коментар"]
    hr = 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers))

    r = hr + 1
    for row in classification:
        excluded = "ВИКЛЮЧЕНО" in (row.get("category") or "")
        ws.cell(row=r, column=1, value=row["domain"]).font = BOLD_FONT if row["domain"] == client_domain else NORMAL_FONT
        ws.cell(row=r, column=2, value=row.get("top10_count")).font = NORMAL_FONT
        ws.cell(row=r, column=3, value=row.get("category")).font = NORMAL_FONT
        ws.cell(row=r, column=4, value=row.get("comment")).font = NORMAL_FONT
        fill = CLIENT_FILL if row["domain"] == client_domain else (EXCLUDED_FILL if excluded else None)
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
        r += 1
    autofit(ws, [20, 14, 26, 55])
    ws.freeze_panes = "A4"


def add_metrics_sheet(wb, metrics, client_domain, excluded_domains):
    ws = wb.active
    ws.title = "SEO показники"
    ws["A1"] = "SEO показники (DataForSEO Labs)"
    ws["A1"].font = TITLE_FONT
    note = ("Сторінок в індексі = сторінки, що ранжуються хоча б по 1 ключовому слову (Relevant Pages) · "
            "Домени-донори = referring domains (Backlinks Summary) · "
            "Ключові слова = органічні ключові слова (Domain Rank Overview) · "
            "SEO трафік = оцінка органічного трафіку/міс (ETV).")
    if excluded_domains:
        note += " Виключено як нерелевантні: " + ", ".join(excluded_domains) + "."
    ws["A2"] = note
    ws["A2"].font = SUB_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 40

    rows = [m for m in metrics if m["domain"] not in excluded_domains]

    headers = ["Сайт", "Сторінок в індексі", "Домени-донори", "Ключові слова", "SEO трафік (міс.)"]
    hr = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers))

    r = hr + 1
    first_data_row = r
    for m in rows:
        ws.cell(row=r, column=1, value=m["domain"]).font = BOLD_FONT if m["domain"] == client_domain else NORMAL_FONT
        for c, key in zip(range(2, 6), ["pages", "referring_domains", "keywords", "traffic"]):
            cell = ws.cell(row=r, column=c, value=m.get(key))
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
            cell.font = NORMAL_FONT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
        if m["domain"] == client_domain:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = CLIENT_FILL
        r += 1
    last_data_row = r - 1

    competitor_rows = [m for m in rows if m["domain"] != client_domain]

    def summary_row(label, fn):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).border = BORDER
        for c, key in zip(range(2, 6), ["pages", "referring_domains", "keywords", "traffic"]):
            vals = [m[key] for m in competitor_rows if m.get(key) is not None]
            val = fn(vals) if vals else None
            cell = ws.cell(row=r, column=c, value=val)
            cell.number_format = "#,##0"
            cell.font = BOLD_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
        r += 1

    summary_row("Середнє по конкурентах (без клієнта)*", lambda v: round(sum(v) / len(v)))
    summary_row("Максимум по конкурентах (без клієнта)*", max)

    ws.cell(row=r, column=1, value="* статичний знімок на дату збору даних, не жива формула — див. коментар у build_report.py")
    ws.cell(row=r, column=1).font = SUB_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

    autofit(ws, [22, 18, 16, 16, 18])
    ws.freeze_panes = "A5"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--serp", help="serp_results.json from serp_top10.py")
    ap.add_argument("--classification", help="hand-authored classification.json (see this script's docstring)")
    ap.add_argument("--metrics", required=True, help="metrics.json from domain_metrics.py")
    ap.add_argument("--client", required=True, help="the client's own domain, to highlight it")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    metrics = json.loads(Path(args.metrics).read_text(encoding="utf-8"))

    excluded_domains = set()
    classification = None
    if args.classification:
        classification = json.loads(Path(args.classification).read_text(encoding="utf-8"))
        excluded_domains = {row["domain"] for row in classification if "ВИКЛЮЧЕНО" in (row.get("category") or "")}

    wb = Workbook()
    # metrics sheet claims the default first sheet; add others after, then
    # reorder so SERP/classification (if present) read first, metrics last —
    # matches the natural "gather -> classify -> compare" reading order.
    add_metrics_sheet(wb, metrics, args.client, excluded_domains)
    metrics_ws = wb["SEO показники"]

    if args.serp:
        serp_rows = json.loads(Path(args.serp).read_text(encoding="utf-8"))
        add_serp_sheet(wb, serp_rows)
    if classification:
        add_classification_sheet(wb, classification, args.client)

    wb.move_sheet("SEO показники", offset=len(wb.sheetnames) - 1)

    wb.save(args.out)
    print(f"Saved workbook -> {args.out}")


if __name__ == "__main__":
    main()
